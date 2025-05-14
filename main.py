import os
import shutil
from datetime import datetime, timedelta
from src import fiserv_parser, excel_generator, mailer
from openpyxl.styles import PatternFill
import openpyxl
import argparse
from dotenv import load_dotenv
import json
import hashlib
import sys
import base64
from cryptography.fernet import Fernet

# Definir rutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(BASE_DIR, 'config', '.env')
INPUT_DIR = os.path.join(BASE_DIR, 'files', 'input')
TERMINADO_DIR = os.path.join(BASE_DIR, 'files', 'terminado')
EXCEL_DIR = os.path.join(BASE_DIR, 'files', 'excel')
LICENSE_FILE = os.path.join(BASE_DIR, 'config', 'fiserv_conciliacion.lic')

# Verificar que el archivo .env existe
if not os.path.exists(ENV_FILE):
    print(f"Error: No se encuentra el archivo .env en {ENV_FILE}")
    sys.exit(1)

# Cargar variables de entorno
load_dotenv(ENV_FILE)

# Verificar carga de variables críticas
def verify_env_vars():
    required_vars = ['SMTP_SERVER', 'SMTP_PORT', 'SMTP_USER', 'SMTP_PASSWORD', 'MAIL_RECIPIENTS']
    missing_vars = []
    for var in required_vars:
        value = os.getenv(var)
        if value is None or value.strip() == '':
            missing_vars.append(var)
    
    if missing_vars:
        print("Error: Faltan las siguientes variables en el archivo .env:")
        print(f"Archivo .env ubicado en: {ENV_FILE}")
        for var in missing_vars:
            print(f"- {var}")
            # Mostrar el valor actual (para debug)
            print(f"  Valor actual: '{os.getenv(var, '')}'")
        sys.exit(1)
    print("Variables de entorno cargadas correctamente")

verify_env_vars()

os.makedirs(TERMINADO_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

def generate_encryption_key():
    """Genera una clave de encriptación única basada en el hardware"""
    system_info = os.getenv('COMPUTERNAME', '') + os.getenv('USERNAME', '')
    key = hashlib.sha256(system_info.encode()).digest()
    return base64.urlsafe_b64encode(key)

def verify_license():
    try:
        if not os.path.exists(LICENSE_FILE):
            print("Error: Archivo de licencia no encontrado")
            return False
        
        # Leer y desencriptar la licencia
        with open(LICENSE_FILE, 'rb') as f:
            encrypted_data = f.read()
        
        try:
            f = Fernet(generate_encryption_key())
            decrypted_data = f.decrypt(encrypted_data)
            license_data = json.loads(decrypted_data)
        except Exception as e:
            print("Error: Licencia inválida o corrupta")
            return False
        
        # Verificar que la licencia corresponde a este equipo
        if license_data['system_id'] != os.getenv('COMPUTERNAME', ''):
            print("Error: Esta licencia no corresponde a este equipo")
            return False
        
        # Verificar fecha de vencimiento
        expiration_date = datetime.strptime(license_data['expiration_date'], '%Y-%m-%d')
        if datetime.now() > expiration_date:
            print(f"Error: Licencia vencida. Fecha de vencimiento: {expiration_date.strftime('%Y-%m-%d')}")
            return False
        
        # Verificar hash de la licencia
        license_key = license_data['license_key']
        expected_hash = hashlib.sha256(license_key.encode()).hexdigest()
        if license_data['hash'] != expected_hash:
            print("Error: Licencia inválida")
            return False
        
        return True
    except Exception as e:
        print(f"Error al verificar licencia: {str(e)}")
        return False

def extract_date_from_filename(filename):
    # Extrae la fecha del formato TRONA001.20250403...
    try:
        parts = filename.split('.')
        if len(parts) < 2:
            return None
        date_str = parts[1][:8]
        if len(date_str) != 8:
            return None
        date_obj = datetime.strptime(date_str, '%Y%m%d')
        return date_obj
    except (IndexError, ValueError) as e:
        print(f"Error al procesar fecha: {e}")
        return None

def get_args_env():
    parser = argparse.ArgumentParser()
    parser.add_argument('--frecuencia', choices=['diario', 'semanal', 'mensual'], help='Frecuencia de envío')
    parser.add_argument('--dia-inicio-semana', type=int, choices=range(0,7), help='Día de inicio de semana (0=Lunes, 6=Domingo)')
    parser.add_argument('--semana-anterior', action='store_true', help='Procesar la semana anterior en lugar de la actual')
    parser.add_argument('--year', type=int, help='Año para procesamiento mensual')
    parser.add_argument('--month', type=int, help='Mes para procesamiento mensual')
    args = parser.parse_args()
    # Leer de env si no está en args
    frecuencia = args.frecuencia or os.getenv('FRECUENCIA_ENVIO', 'diario')
    dia_inicio_semana = args.dia_inicio_semana
    if dia_inicio_semana is None:
        try:
            dia_inicio_semana = int(os.getenv('DIA_INICIO_SEMANA', '0'))
        except Exception:
            dia_inicio_semana = 0
    semana_anterior = args.semana_anterior
    year = args.year
    month = args.month
    return frecuencia, dia_inicio_semana, semana_anterior, year, month

def archivos_por_frecuencia(frecuencia, dia_inicio_semana, semana_anterior=False, year=None, month=None):
    hoy = datetime.now()
    archivos = []
    fechas = []
    for filename in os.listdir(INPUT_DIR):
        if 'CL586D' not in filename or not filename.endswith('.TXT'):
            continue
        fecha = extract_date_from_filename(filename)
        if not fecha:
            continue
        fechas.append(fecha)
        archivos.append((filename, fecha))
    seleccionados = []
    if frecuencia == 'diario':
        for fname, fdate in archivos:
            if fdate.date() == hoy.date():
                seleccionados.append(fname)
    elif frecuencia == 'semanal':
        # Calcular inicio de semana configurable
        dias_desde_inicio = (hoy.weekday() - dia_inicio_semana) % 7
        inicio_semana = hoy - timedelta(days=dias_desde_inicio)
        if semana_anterior:
            # Retroceder 7 días para obtener la semana anterior
            inicio_semana = inicio_semana - timedelta(days=7)
            fin_semana = inicio_semana + timedelta(days=6)
        else:
            fin_semana = hoy
        inicio_semana = inicio_semana.replace(hour=0, minute=0, second=0, microsecond=0)
        for fname, fdate in archivos:
            if inicio_semana.date() <= fdate.date() <= fin_semana.date():
                seleccionados.append(fname)
    elif frecuencia == 'mensual':
        if year is not None and month is not None:
            # Procesar año y mes específicos
            target_year = year
            target_month = month
        else:
            # Procesar mes anterior por defecto
            primer_dia_mes_actual = hoy.replace(day=1)
            ultimo_mes = primer_dia_mes_actual - timedelta(days=1)
            target_year = ultimo_mes.year
            target_month = ultimo_mes.month
            print(f"Procesando mes anterior: {target_month}/{target_year}")
        
        for fname, fdate in archivos:
            if fdate.year == target_year and fdate.month == target_month:
                seleccionados.append(fname)
    return seleccionados

def main():
    # Verificar licencia
    if not verify_license():
        sys.exit(1)
        
    frecuencia, dia_inicio_semana, semana_anterior, year, month = get_args_env()
    print(f"Frecuencia: {frecuencia}, Día inicio semana: {dia_inicio_semana}, Semana anterior: {semana_anterior}")
    if year and month:
        print(f"Procesando año: {year}, mes: {month}")
    
    seleccionados = archivos_por_frecuencia(frecuencia, dia_inicio_semana, semana_anterior, year, month)
    if not seleccionados:
        print("No hay archivos para procesar en el rango seleccionado.")
        return
    print(f"Archivos a procesar: {seleccionados}")

    # Unir todos los registros de todos los archivos
    registros_unificados = {k: [] for k in ['1','2','3','6','7','8','9','unknown']}
    fechas_archivos = []
    for filename in seleccionados:
        txt_path = os.path.join(INPUT_DIR, filename)
        fecha_archivo = extract_date_from_filename(filename)
        if fecha_archivo:
            fechas_archivos.append(fecha_archivo)
        print(f"Procesando: {filename}")
        records = fiserv_parser.parse_txt(txt_path)
        for k in registros_unificados:
            registros_unificados[k].extend(records.get(k, []))
        shutil.move(txt_path, os.path.join(TERMINADO_DIR, filename))
        print(f"Procesado: {filename}")

    # Determinar rango de fechas para el nombre del Excel y asunto
    if fechas_archivos:
        fecha_min = min(fechas_archivos).strftime('%d-%m-%Y')
        fecha_max = max(fechas_archivos).strftime('%d-%m-%Y')
        rango = f"{fecha_min}_a_{fecha_max}"
    else:
        rango = datetime.now().strftime('%d-%m-%Y')

    excel_name = f"Conciliacion_{frecuencia}_{rango}.xlsx"
    excel_path = excel_generator.generate_excel(registros_unificados, excel_name, EXCEL_DIR)
    mailer.send_files([excel_path], rango)
    print(f"Procesado y enviado: {excel_name}")

    # Aplicar estilos después de ajustar anchos de columna
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    wb = openpyxl.load_workbook(excel_path)
    for ws in wb.worksheets:
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            cell.fill = header_fill
    wb.save(excel_path)

if __name__ == "__main__":
    main()
